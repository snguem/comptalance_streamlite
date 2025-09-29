
from datetime import datetime
import os
import streamlit as st
import io
import json
from openpyxl import workbook, load_workbook
from PIL import Image


def load_models():
    models = []
    try:
        with open("models.json") as m:
            models = json.load(m)
    except:
        pass
    
    return models


class Config:
    def __init__(self):
        self.model_balances_ = {"modele":False, "baln": False, "baln_1":False}
        self.isIntegrated = False
        self.model_wb = None
        self.baln_wb = None
        self.baln_1_wb = None
        self.excel_data = None
        self.modeles = load_models()
        self.default_img_path = "models_images"
        self.default_excel_folder = "models_excel"
    
    
    def _copy_worksheet_with_styles(self,st, source_ws, target_ws):
        """Copie une feuille Excel en préservant tous les styles et formatages"""
        try:
            # Copie des dimensions des colonnes
            for col_letter, dimension in source_ws.column_dimensions.items():
                target_ws.column_dimensions[col_letter].width = dimension.width
                target_ws.column_dimensions[col_letter].hidden = dimension.hidden
            
            # Copie des dimensions des lignes
            for row_num, dimension in source_ws.row_dimensions.items():
                target_ws.row_dimensions[row_num].height = dimension.height
                target_ws.row_dimensions[row_num].hidden = dimension.hidden
            
            # Copie des cellules avec tous leurs attributs
            for row in source_ws.iter_rows():
                for cell in row:
                    target_cell = target_ws.cell(row=cell.row, column=cell.column)
                    
                    # Copie de la valeur
                    target_cell.value = cell.value
                    
                    # Copie du style si disponible
                    if hasattr(cell, 'has_style') and cell.has_style:
                        if cell.font:
                            target_cell.font = cell.font.copy()
                        if cell.fill:
                            target_cell.fill = cell.fill.copy()
                        if cell.border:
                            target_cell.border = cell.border.copy()
                        if cell.alignment:
                            target_cell.alignment = cell.alignment.copy()
                        if cell.number_format:
                            target_cell.number_format = cell.number_format
                        if cell.protection:
                            target_cell.protection = cell.protection.copy()
            
            # Copie des cellules fusionnées
            for merged_range in source_ws.merged_cells.ranges:
                target_ws.merge_cells(str(merged_range))
            
            # Copie des paramètres de mise en page (avec vérification d'existence)
            try:
                if hasattr(source_ws, 'page_setup') and source_ws.page_setup:
                    target_ws.page_setup = source_ws.page_setup
            except:
                pass
            
            try:
                if hasattr(source_ws, 'page_margins') and source_ws.page_margins:
                    target_ws.page_margins = source_ws.page_margins
            except:
                pass
            
            try:
                if hasattr(source_ws, 'header_footer') and source_ws.header_footer:
                    target_ws.header_footer = source_ws.header_footer
            except:
                pass
            
            # Copie des paramètres d'impression additionnels
            try:
                if hasattr(source_ws, 'print_options') and source_ws.print_options:
                    target_ws.print_options = source_ws.print_options
            except:
                pass
                
            try:
                if hasattr(source_ws, 'sheet_view') and source_ws.sheet_view:
                    target_ws.sheet_view = source_ws.sheet_view
            except:
                pass
            
        except Exception as e:
            pass
            # st.warning(f"Copie avec styles partielle: {str(e)}. Basculement vers copie simple.")
            # Fallback vers copie simple
            # self._copy_worksheet_simple(source_ws, target_ws)
    
    
    def add_balances_to_modele(self, st):
        """Ajoute les balances au modèle Excel en préservant leur style original"""
        try:
            # Suppression des anciennes feuilles si elles existent
            if 'BAL N' in self.modele_wb.sheetnames:
                del self.modele_wb['BAL N']
            if 'BAL N-1' in self.modele_wb.sheetnames:
                del self.modele_wb['BAL N-1']
            
            # Ajout de la balance N en préservant le style original
            if self.baln_wb is not None:
                # Copie de la feuille originale avec tous ses styles
                source_sheet = self.baln_wb.active
                target_sheet = self.modele_wb.create_sheet('BAL N')
                self._copy_worksheet_with_styles(st, source_sheet, target_sheet)
            else:
                pass
            
            # Ajout de la balance N-1 si disponible
            if self.baln_1_wb is not None:
                # Copie de la feuille originale avec tous ses styles
                source_sheet = self.baln_1_wb.active
                target_sheet = self.modele_wb.create_sheet('BAL N-1')
                self._copy_worksheet_with_styles(st, source_sheet, target_sheet)
            else:
                pass
            
            return True
        except Exception as e:
            st.error(f"Erreur lors de l'integration des balances: {str(e)}")
            return False
    
    
    def get_excel_file_download(self, st):
        """Génère le fichier Excel téléchargeable"""
        if not self.modele_wb:
            return None
        
        try:
            buffer = io.BytesIO()
            self.modele_wb.save(buffer)
            buffer.seek(0)
            self.excel_data= buffer.getvalue()
            return self.excel_data
        except Exception as e:
            st.error(f"Erreur lors de la génération du fichier: {str(e)}")
            return None
    
    
    def display_image(self, st, image_path):
        """Affiche une image avec gestion d'erreur"""
        try:
            if os.path.exists(image_path):
                image = Image.open(image_path)
                st.image(image, use_container_width=True)
                # return True
            else:
                st.warning(f"Image non trouvée: {image_path}")
                # return False
        except Exception as e:
            st.error(f"Erreur lors du chargement de l'image: {str(e)}")
            # return False
    
    def get_index_model(self, model_choisi:str):
        if self.modeles:
            for i, model in enumerate(self.modeles):
                if model["nom"].lower() == model_choisi.lower():
                    return i
        
        return 0

    def load_excel(self, st, uploaded_file, type=1):
        try:
            if type==1:
                self.modele_wb = load_workbook(uploaded_file, data_only=False)
            elif type==2:
                self.baln_wb = load_workbook(uploaded_file, data_only=False)
            else:
                self.baln_1_wb = load_workbook(uploaded_file, data_only=False)
            return True
        except Exception as e:
            st.error(f"Erreur lors du chargement du modèle: {str(e)}")
            return False
    
    def model_choisi(self):
        self.model_balances_["modele"] = True
        
    def balance_n_importer(self):
        self.model_balances_["baln"] = True
        
    def balance_n_1_importer(self):
        self.model_balances_["baln_1"] = True



# global variable

# Initialisation de l'application
if 'config' not in st.session_state:
    st.session_state.config = Config()
    
config = st.session_state.config
    
if 'current_model_index' not in st.session_state:
    st.session_state.current_model_index = 0

if config.load_excel(st, os.path.join(config.default_excel_folder, config.modeles[0]["file_path"])):
    config.model_choisi()

# Configuration de la page
st.set_page_config(
    page_title="Logiciel d'integration de balances comptable",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personnalisé
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
        font-weight: bold;
    }
    .section-header {
        font-size: 1.5rem;
        color: #ff7f0e;
        margin: 1rem 0;
        border-bottom: 2px solid #ff7f0e;
        padding-bottom: 0.5rem;
    }
    .success-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        margin: 1rem 0;
    }
    .info-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        margin: 1rem 0;
    }
    .warning-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)


st.markdown('<h1 class="main-header">📊 Comptalance, le logiciel qui facilite l\'integration de vos balance N et N-1</h1>', unsafe_allow_html=True)

# fonctions

# =====================

# Sidebar pour la navigation
with st.sidebar:
    st.markdown("### Navigation")
    page = st.selectbox(
        "Choisir une section:",
        ["🏠 Accueil", "📁 Choisir un model", "📤 Importer les balances"]
    )
    
    st.markdown("---")
    st.markdown("### Informations")
    st.info("Logiciel développé pour l'integration rapide des balances N et N-1")


if page == "🏠 Accueil":
    st.markdown('<div class="section-header">Bienvenue</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        ### 💡 Attention
        
        - Les balances **ne doivent pas** etre dans le meme fichier excel
        - Il ne doit avoir qu'une page(balance N ou N-1) dans le fichier que vous importer
        - Vous devez avoir 2 fichiers excel : celui de la balance N et N-1
        """)
    
    with col2:
        st.markdown("""
        ### 📋 Guide d'utilisation
        
        1. Aller a la section **choisir un modele** en haut a gauche
        2. Choisissez le modele que vous souhaitez
        3. Aller a la section **Importer les Balances** en haut a gauche
        4. **Ajoutez** vos balances (2 fichiers excels)
        5. Revenez a l'acceuil et scroller jusqu'a apercevoir l'etat des fichiers
        6. **Lancer l'integration** et **Exporter le resultat au format excel**
        """)
    
    # Statut des fichiers
    st.markdown('<div class="section-header">État des fichiers</div>', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        modele_status = "✅ Chargé" if config.model_balances_["modele"] else "❌ Non choisi"
        st.metric("Modèle Excel", modele_status)
    
    with col2:
        balance_n_status = "✅ Chargée" if config.model_balances_["baln"] else "❌ Non chargée"
        st.metric("Balance N", balance_n_status)
    
    with col3:
        balance_n1_status = "✅ Chargée" if config.model_balances_["baln_1"] else "❌ Non chargée"
        st.metric("Balance N-1", balance_n1_status)


    if config.model_balances_["modele"] and config.model_balances_["baln"] and config.model_balances_["baln_1"]:
        st.markdown('<div class="section-header">Integration</div>', unsafe_allow_html=True)
        # Bouton pour intégrer les balances
        
        if st.button("🔄 Intégrer les balances au modèle", type="primary"):
            with st.spinner("Traitement en cours..."):
                if config.add_balances_to_modele(st):
                    config.isIntegrated = True
                    config.get_excel_file_download(st)
                    st.info("Les feuilles 'Balance_N' et 'Balance_N-1' ont été ajoutées au modèle")
                else:
                    st.markdown('<div class="warning-box">❌ Erreur lors de l\'intégration des balances</div>', unsafe_allow_html=True)
            # st.success("Terminé!")

    if config.isIntegrated:
        st.markdown('<div class="section-header">Exportation</div>', unsafe_allow_html=True)
        # Bouton pour intégrer les balances
        st.download_button(
            label="📥 Télécharger Excel",
            data=config.excel_data,
            file_name=f"comptabilite_complete_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        
        
elif page == "📁 Choisir un model":
    # Interface principale
    st.header("📋 Liste des Modèles Disponibles")
    
    # Création de deux colonnes
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.subheader("Modeles disponible")
        
        # Liste des modèles avec sélection
        model_names = [model['nom'] for model in config.modeles]
        
        if 'current_model_index' not in st.session_state:
            st.session_state.current_model_index = 0
        
        
        model_choisi = st.radio(
            "djf",
            model_names,
            index=0,
            label_visibility="hidden"
        )
        
        index = config.get_index_model(model_choisi)
        if index!=st.session_state.current_model_index:
            st.session_state.current_model_index = index
            if config.load_excel(st, os.path.join(config.default_excel_folder, config.modeles[index]["file_path"])):
                config.model_choisi()
        
    with col2:
        if config.modeles:
            selected_model = config.modeles[st.session_state.current_model_index]
            
            st.subheader(f"🖼️ Aparcu du {selected_model['nom']}")
            
            # Affichage des images
            if 'imgs' in selected_model and selected_model['imgs']:
                # st.markdown()
                
                # Options d'affichage des images
                show_images = st.checkbox("Afficher les images", value=False)
                
                if show_images:
                    for i, image_name in enumerate(selected_model['imgs']):
                        image_path = os.path.join(config.default_img_path, image_name)
                        
                        st.markdown(f"**Image {i+1}:** {image_name}")
                        
                        config.display_image(st, image_path)
                        st.markdown("---")
            else:
                st.info("Aucune image disponible pour ce modèle")
    
    st.markdown("---")
    

elif page == "📤 Importer les balances":
    config.model_balances_["baln"] = False
    config.model_balances_["baln_1"] = False
    st.markdown('<div class="section-header">Import des balances</div>', unsafe_allow_html=True)
    
    # Upload des balances
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 📊 Balance Année N")
        balance_n_file = st.file_uploader(
            "Fichier balance année courante (N):",
            type=['xlsx'],
            key="balance_n_uploader"
        )
        
        if balance_n_file:
            if config.load_excel(st, balance_n_file, 2):
                config.model_balances_["baln"] = True
                st.success("✅ Balance N chargée!")
    
    with col2:
        st.markdown("#### 📊 Balance Année N-1")
        balance_n1_file = st.file_uploader(
            "Fichier balance année précédente (N-1):",
            type=['xlsx'],
            key="balance_n1_uploader"
        )
        
        if balance_n1_file:
            if config.load_excel(st, balance_n1_file, 3):
                config.model_balances_["baln_1"] = True
                st.success("✅ Balance N-1 chargée!")
    
    st.markdown("---")
    

if __name__=="__main__":
    pass