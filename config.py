
import io
import json
import os
from openpyxl import workbook, load_workbook
from PIL import Image


def load_models():
    models = []
    try:
        with open("models.json") as m:
            models = json.load(m)
    except:
        pass
    
    return models


class Config:
    def __init__(self):
        self.model_balances_ = {"modele":False, "baln": False, "baln_1":False}
        self.isIntegrated = False
        self.model_wb = None
        self.baln_wb = None
        self.baln_1_wb = None
        self.excel_data = None
        self.modeles = load_models()
        self.default_img_path = "models_images"
        self.default_excel_folder = "models_excel"
    
    
    def _copy_worksheet_with_styles(self,st, source_ws, target_ws):
        """Copie une feuille Excel en préservant tous les styles et formatages"""
        try:
            # Copie des dimensions des colonnes
            for col_letter, dimension in source_ws.column_dimensions.items():
                target_ws.column_dimensions[col_letter].width = dimension.width
                target_ws.column_dimensions[col_letter].hidden = dimension.hidden
            
            # Copie des dimensions des lignes
            for row_num, dimension in source_ws.row_dimensions.items():
                target_ws.row_dimensions[row_num].height = dimension.height
                target_ws.row_dimensions[row_num].hidden = dimension.hidden
            
            # Copie des cellules avec tous leurs attributs
            for row in source_ws.iter_rows():
                for cell in row:
                    target_cell = target_ws.cell(row=cell.row, column=cell.column)
                    
                    # Copie de la valeur
                    target_cell.value = cell.value
                    
                    # Copie du style si disponible
                    if hasattr(cell, 'has_style') and cell.has_style:
                        if cell.font:
                            target_cell.font = cell.font.copy()
                        if cell.fill:
                            target_cell.fill = cell.fill.copy()
                        if cell.border:
                            target_cell.border = cell.border.copy()
                        if cell.alignment:
                            target_cell.alignment = cell.alignment.copy()
                        if cell.number_format:
                            target_cell.number_format = cell.number_format
                        if cell.protection:
                            target_cell.protection = cell.protection.copy()
            
            # Copie des cellules fusionnées
            for merged_range in source_ws.merged_cells.ranges:
                target_ws.merge_cells(str(merged_range))
            
            # Copie des paramètres de mise en page (avec vérification d'existence)
            try:
                if hasattr(source_ws, 'page_setup') and source_ws.page_setup:
                    target_ws.page_setup = source_ws.page_setup
            except:
                pass
            
            try:
                if hasattr(source_ws, 'page_margins') and source_ws.page_margins:
                    target_ws.page_margins = source_ws.page_margins
            except:
                pass
            
            try:
                if hasattr(source_ws, 'header_footer') and source_ws.header_footer:
                    target_ws.header_footer = source_ws.header_footer
            except:
                pass
            
            # Copie des paramètres d'impression additionnels
            try:
                if hasattr(source_ws, 'print_options') and source_ws.print_options:
                    target_ws.print_options = source_ws.print_options
            except:
                pass
                
            try:
                if hasattr(source_ws, 'sheet_view') and source_ws.sheet_view:
                    target_ws.sheet_view = source_ws.sheet_view
            except:
                pass
            
        except Exception as e:
            pass
            # st.warning(f"Copie avec styles partielle: {str(e)}. Basculement vers copie simple.")
            # Fallback vers copie simple
            # self._copy_worksheet_simple(source_ws, target_ws)
    
    
    def add_balances_to_modele(self, st):
        """Ajoute les balances au modèle Excel en préservant leur style original"""
        try:
            # Suppression des anciennes feuilles si elles existent
            if 'BAL N' in self.modele_wb.sheetnames:
                del self.modele_wb['BAL N']
            if 'BAL N-1' in self.modele_wb.sheetnames:
                del self.modele_wb['BAL N-1']
            
            # Ajout de la balance N en préservant le style original
            if self.baln_wb is not None:
                # Copie de la feuille originale avec tous ses styles
                source_sheet = self.baln_wb.active
                target_sheet = self.modele_wb.create_sheet('BAL N')
                self._copy_worksheet_with_styles(st, source_sheet, target_sheet)
            else:
                pass
            
            # Ajout de la balance N-1 si disponible
            if self.baln_1_wb is not None:
                # Copie de la feuille originale avec tous ses styles
                source_sheet = self.baln_1_wb.active
                target_sheet = self.modele_wb.create_sheet('BAL N-1')
                self._copy_worksheet_with_styles(st, source_sheet, target_sheet)
            else:
                pass
            
            return True
        except Exception as e:
            st.error(f"Erreur lors de l'integration des balances: {str(e)}")
            return False
    
    
    def get_excel_file_download(self, st):
        """Génère le fichier Excel téléchargeable"""
        if not self.modele_wb:
            return None
        
        try:
            buffer = io.BytesIO()
            self.modele_wb.save(buffer)
            buffer.seek(0)
            self.excel_data= buffer.getvalue()
            return self.excel_data
        except Exception as e:
            st.error(f"Erreur lors de la génération du fichier: {str(e)}")
            return None
    
    
    def display_image(self, st, image_path):
        """Affiche une image avec gestion d'erreur"""
        try:
            if os.path.exists(image_path):
                image = Image.open(image_path)
                st.image(image, use_container_width=True)
                # return True
            else:
                st.warning(f"Image non trouvée: {image_path}")
                # return False
        except Exception as e:
            st.error(f"Erreur lors du chargement de l'image: {str(e)}")
            # return False
    
    def get_index_model(self, model_choisi:str):
        if self.modeles:
            for i, model in enumerate(self.modeles):
                if model["nom"].lower() == model_choisi.lower():
                    return i
        
        return 0

    def load_excel(self, st, uploaded_file, type=1):
        try:
            if type==1:
                self.modele_wb = load_workbook(uploaded_file, data_only=False)
            elif type==2:
                self.baln_wb = load_workbook(uploaded_file, data_only=False)
            else:
                self.baln_1_wb = load_workbook(uploaded_file, data_only=False)
            return True
        except Exception as e:
            st.error(f"Erreur lors du chargement du modèle: {str(e)}")
            return False
    
    def model_choisi(self):
        self.model_balances_["modele"] = True
        
    def balance_n_importer(self):
        self.model_balances_["baln"] = True
        
    def balance_n_1_importer(self):
        self.model_balances_["baln_1"] = True


if __name__=="__main__":
    print(load_models())

